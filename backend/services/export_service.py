import pandas as pd
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from backend.services.dashboard_service import (
    get_dashboard_stats,
    get_latest_screenings_per_patient,
)


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_candidates_excel(db: Session, trial_id: str) -> bytes:
    """
    Export screened candidates as a professional Excel (.xlsx) file.
    """

    screenings = get_latest_screenings_per_patient(db, trial_id)

    data = []

    for s in screenings:

        patient = getattr(s, "patient", None)

        patient_name = patient.name if patient else "Unknown"
        patient_phone = patient.phone if patient else "N/A"

        data.append({
            "Patient ID": s.patient_id,
            "Patient Name": patient_name,
            "Contact Phone": patient_phone,
            "Match Percentage": (
                float(s.match_percentage)
                if s.match_percentage is not None
                else 0
            ),
            "Verdict": s.verdict or "N/A",
            "Eligible": "Yes" if s.eligible else "No",
            "Screened At": (
                s.screened_at.strftime("%Y-%m-%d %H:%M")
                if s.screened_at
                else "N/A"
            ),
        })

    columns = [
        "Patient ID",
        "Patient Name",
        "Contact Phone",
        "Match Percentage",
        "Verdict",
        "Eligible",
        "Screened At",
    ]

    df = pd.DataFrame(data, columns=columns)

    excel_buffer = BytesIO()

    with pd.ExcelWriter(
        excel_buffer,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Candidates",
            index=False,
            startrow=2,
        )

        workbook = writer.book
        worksheet = writer.sheets["Candidates"]

        # ----------------------------------------------------
        # Imports for Excel formatting
        # ----------------------------------------------------
        from openpyxl.styles import (
            Font,
            PatternFill,
            Alignment,
            Border,
            Side,
        )
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # ----------------------------------------------------
        # Title
        # ----------------------------------------------------

        worksheet["A1"] = f"Clinical Trial Candidate Report - {trial_id}"
        worksheet["A1"].font = Font(
            bold=True,
            size=16
        )

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(columns)
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center"
        )

        # ----------------------------------------------------
        # Header formatting
        # ----------------------------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        thin_border = Border(
            bottom=Side(
                style="thin",
                color="D9E1F2"
            )
        )

        for cell in worksheet[3]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = thin_border

        # ----------------------------------------------------
        # Percentage formatting
        # ----------------------------------------------------

        percentage_column = 4

        for row in range(4, worksheet.max_row + 1):

            cell = worksheet.cell(
                row=row,
                column=percentage_column
            )

            # Convert 90 -> 90%
            cell.value = cell.value / 100
            cell.number_format = "0.0%"

        # ----------------------------------------------------
        # Center important columns
        # ----------------------------------------------------

        for row in worksheet.iter_rows(
            min_row=4,
            max_row=worksheet.max_row
        ):

            row[0].alignment = Alignment(
                horizontal="center"
            )

            row[2].alignment = Alignment(
                horizontal="center"
            )

            row[3].alignment = Alignment(
                horizontal="center"
            )

            row[4].alignment = Alignment(
                horizontal="center"
            )

            row[5].alignment = Alignment(
                horizontal="center"
            )

            row[6].alignment = Alignment(
                horizontal="center"
            )

        # ----------------------------------------------------
        # Excel table
        # ----------------------------------------------------

        if worksheet.max_row >= 4:

            table_ref = (
                f"A3:G{worksheet.max_row}"
            )

            table = Table(
                displayName="CandidateTable",
                ref=table_ref
            )

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style

            worksheet.add_table(table)

        # ----------------------------------------------------
        # Freeze header
        # ----------------------------------------------------

        worksheet.freeze_panes = "A4"

        # ----------------------------------------------------
        # Auto column width
        # ----------------------------------------------------
        from openpyxl.utils import get_column_letter

        for column_index in range(1, worksheet.max_column + 1):

            max_length = 0

            column_letter = get_column_letter(column_index)

            for row in range(1, worksheet.max_row + 1):

                cell = worksheet.cell(
                    row=row,
                    column=column_index
                )

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                35
            )
        

        # ----------------------------------------------------
        # Row heights
        # ----------------------------------------------------

        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[3].height = 24

    excel_buffer.seek(0)

    return excel_buffer.getvalue()


# ============================================================
# PDF EXPORT
# ============================================================

def export_dashboard_pdf(db: Session, trial_id: str) -> bytes:
    """
    Export a professional clinical trial recruitment dashboard PDF.
    """

    stats = get_dashboard_stats(db, trial_id)

    pdf_buffer = BytesIO()

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=45,
        leftMargin=45,
        topMargin=45,
        bottomMargin=45,
    )

    styles = getSampleStyleSheet()

    # --------------------------------------------------------
    # Custom styles
    # --------------------------------------------------------

    title_style = ParagraphStyle(
        "TitleCustom",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20,
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        spaceBefore=15,
        spaceAfter=10,
    )

    normal_style = ParagraphStyle(
        "NormalCustom",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
    )

    # --------------------------------------------------------
    # Story
    # --------------------------------------------------------

    story = []

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Clinical Trial Recruitment Report",
            title_style
        )
    )

    story.append(
        Paragraph(
            f"<b>Trial ID:</b> {trial_id}",
            subtitle_style
        )
    )

    story.append(Spacer(1, 5))

    # --------------------------------------------------------
    # Recruitment Summary
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Recruitment Summary",
            section_style
        )
    )

    target = stats.get("target", 0)
    enrolled = stats.get("enrolled", 0)
    screened = stats.get("screened", 0)
    progress = stats.get("progress", 0.0)

    summary_data = [
        [
            "Target Recruitment",
            "Currently Enrolled",
            "Total Screened",
            "Progress",
        ],
        [
            str(target),
            str(enrolled),
            str(screened),
            f"{progress:.1f}%",
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            1.7 * inch,
            1.7 * inch,
            1.7 * inch,
            1.2 * inch,
        ],
    )

    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF2F8")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C7DC")),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ])
    )

    story.append(summary_table)

    # --------------------------------------------------------
    # Screening Results
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Screening Results",
            section_style
        )
    )

    approved = stats.get("approved", 0)
    needs_review = stats.get("needs_review", 0)
    rejected = stats.get("rejected", 0)

    screening_data = [
        ["Status", "Candidates"],
        ["Approved", str(approved)],
        ["Needs Review", str(needs_review)],
        ["Rejected", str(rejected)],
    ]

    screening_table = Table(
        screening_data,
        colWidths=[
            3.5 * inch,
            2.8 * inch,
        ],
    )

    screening_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C7DC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#F5F8FA"),
            ]),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ])
    )

    story.append(screening_table)

    # --------------------------------------------------------
    # Exclusion Reasons
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Top Exclusion Reasons",
            section_style
        )
    )

    exclusion_reasons = stats.get(
        "top_exclusion_reasons",
        []
    )

    if exclusion_reasons:

        exclusion_data = [
            ["Reason", "Occurrences"]
        ]

        for reason in exclusion_reasons:

            exclusion_data.append([
                str(reason.get("reason", "Unknown")),
                str(reason.get("count", 0)),
            ])

        exclusion_table = Table(
            exclusion_data,
            colWidths=[
                5.2 * inch,
                1.1 * inch,
            ],
            repeatRows=1,
        )

        exclusion_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1F4E78")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "ALIGN",
                    (1, 1),
                    (1, -1),
                    "CENTER"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor("#B4C7DC")
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F5F8FA"),
                    ]
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),
            ])
        )

        story.append(exclusion_table)

    else:

        story.append(
            Paragraph(
                "No exclusion reasons recorded.",
                normal_style
            )
        )

    # --------------------------------------------------------
    # Footer
    # --------------------------------------------------------

    def add_footer(canvas, document):

        canvas.saveState()

        canvas.setFont(
            "Helvetica",
            8
        )

        canvas.setFillColor(
            colors.HexColor("#666666")
        )

        canvas.drawString(
            45,
            25,
            "Clinical Trial Recruitment System"
        )

        canvas.drawRightString(
            letter[0] - 45,
            25,
            f"Page {document.page}"
        )

        canvas.restoreState()

    # --------------------------------------------------------
    # Build PDF
    # --------------------------------------------------------

    doc.build(
        story,
        onFirstPage=add_footer,
        onLaterPages=add_footer,
    )

    pdf_buffer.seek(0)

    return pdf_buffer.getvalue()